from flask import (
    Blueprint,
    jsonify,
    request,
    send_file
)
from openpyxl.drawing.image import Image
from datetime import (
    datetime,
    timedelta
)

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.drawing.image import Image
import os

from models import (
    Alerta,
    Vehiculo,
    Mantenimiento
)

reportes_bp = Blueprint(
    'reportes',
    __name__
)

# =====================================================
# HELPERS FECHAS
# =====================================================

def obtener_rango_fechas():

    tipo = request.args.get(
        'tipo',
        'mensual'
    )

    hoy = datetime.now()

    # =========================================
    # DIARIO
    # =========================================

    if tipo == 'diario':

        inicio = hoy.replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0
        )

        fin = hoy

    # =========================================
    # SEMANAL
    # =========================================

    elif tipo == 'semanal':

        inicio = hoy - timedelta(days=7)

        fin = hoy

    # =========================================
    # MENSUAL
    # =========================================

    elif tipo == 'mensual':

        inicio = hoy - timedelta(days=30)

        fin = hoy

    # =========================================
    # RANGO PERSONALIZADO
    # =========================================

    else:

        fecha_inicio = request.args.get(
            'fecha_inicio'
        )

        fecha_fin = request.args.get(
            'fecha_fin'
        )

        inicio = datetime.strptime(
            fecha_inicio,
            '%Y-%m-%d'
        )

        fin = datetime.strptime(
            fecha_fin,
            '%Y-%m-%d'
        )

    return inicio, fin


# =====================================================
# ESTILOS EXCEL
# =====================================================

header_fill = PatternFill(
    start_color='1E3A8A',
    end_color='1E3A8A',
    fill_type='solid'
)

title_fill = PatternFill(
    start_color='2563EB',
    end_color='2563EB',
    fill_type='solid'
)

white_font = Font(
    color='FFFFFF',
    bold=True
)

title_font = Font(
    color='FFFFFF',
    bold=True,
    size=18
)

subtitle_font = Font(
    color='FFFFFF',
    bold=True,
    size=13
)

center = Alignment(
    horizontal='center',
    vertical='center'
)

thin = Side(
    border_style='thin',
    color='D1D5DB'
)

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)


# =====================================================
# REPORTE ALERTAS JSON
# =====================================================
from datetime import datetime, time
from flask import request, jsonify

@reportes_bp.route('/alertas', methods=['GET'])
def reporte_alertas():

    inicio, fin = obtener_rango_fechas()

    # =========================================
    # 🔥 NORMALIZAR RANGO (CLAVE DEL PROBLEMA)
    # =========================================
    if hasattr(inicio, "date"):
        inicio = datetime.combine(inicio.date(), time.min)
    else:
        inicio = datetime.combine(inicio, time.min)

    if hasattr(fin, "date"):
        fin = datetime.combine(fin.date(), time.max)
    else:
        fin = datetime.combine(fin, time.max)

    categoria = request.args.get('categoria')
    vehiculo_id = request.args.get('vehiculo_id')

    query = Alerta.query.filter(
        Alerta.created_at >= inicio,
        Alerta.created_at <= fin
    )

    # =========================================
    # FILTRO CATEGORÍA
    # =========================================
    if categoria:
        query = query.filter(
            Alerta.tipo == categoria
        )

    # =========================================
    # FILTRO VEHÍCULO
    # =========================================
    if vehiculo_id:
        query = query.filter(
            Alerta.vehiculo_id == vehiculo_id
        )

    alertas = query.all()

    data = []

    for alerta in alertas:

        vehiculo = Vehiculo.query.get(alerta.vehiculo_id)

        data.append({
            'id': alerta.id,
            'vehiculo': vehiculo.placa if vehiculo else None,
            'tipo': alerta.tipo,
            'categoria': alerta.categoria,
            'prioridad': alerta.prioridad,
            'estado': alerta.estado,
            'mensaje': alerta.mensaje,
            'fecha': alerta.created_at.isoformat() if alerta.created_at else None
        })

    return jsonify(data)
# =====================================================
# EXPORTAR ALERTAS EXCEL
# =====================================================

@reportes_bp.route(
    '/alertas/excel',
    methods=['GET']
)
def exportar_alertas_excel():

    inicio, fin = obtener_rango_fechas()

    categoria = request.args.get(
        'categoria'
    )

    vehiculo_id = request.args.get(
        'vehiculo_id'
    )

    query = Alerta.query.filter(
        Alerta.created_at >= inicio,
        Alerta.created_at <= fin
    )

    # =========================================
    # FILTRO CATEGORÍA
    # =========================================

    if categoria:

        query = query.filter(
            Alerta.tipo == categoria
        )

    # =========================================
    # FILTRO VEHÍCULO
    # =========================================

    if vehiculo_id:

        query = query.filter(
            Alerta.vehiculo_id == vehiculo_id
        )

    alertas = query.all()

    wb = Workbook()

    ws = wb.active

    ws.title = 'Alertas'

    # =================================================
    # HEADER EMPRESA
    # =================================================

    ws.merge_cells('A1:G1')

    ws['A1'] = 'TRANSMENA Y CARGA SAS'

    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center

    ws.merge_cells('A2:G2')

    ws['A2'] = 'REPORTE CORPORATIVO DE ALERTAS'

    ws['A2'].font = subtitle_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = center

    ws.merge_cells('A3:G3')

    ws['A3'] = (
        f'Generado: '
        f'{datetime.now().strftime("%Y-%m-%d %H:%M")}'
    )

    ws['A3'].alignment = center

    # =================================================
    # ENCABEZADOS
    # =================================================

    headers = [

        'ID',
        'VEHÍCULO',
        'TIPO',
        'CATEGORÍA',
        'PRIORIDAD',
        'ESTADO',
        'FECHA'
    ]

    row_num = 5

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(
            row=row_num,
            column=col_num
        )

        cell.value = header
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # =================================================
    # DATA
    # =================================================

    current_row = 6

    for alerta in alertas:

        vehiculo = Vehiculo.query.get(
            alerta.vehiculo_id
        )

        values = [

            alerta.id,

            vehiculo.placa
            if vehiculo else 'N/A',

            alerta.tipo,

            alerta.categoria,

            alerta.prioridad,

            alerta.estado,

            alerta.created_at.strftime(
                '%Y-%m-%d %H:%M'
            ) if alerta.created_at else ''
        ]

        for col_num, value in enumerate(values, 1):

            cell = ws.cell(
                row=current_row,
                column=col_num
            )

            cell.value = value
            cell.border = border
            cell.alignment = center

            # =========================================
            # COLOR PRIORIDAD
            # =========================================

            if col_num == 5:

                if value == 'CRITICA':

                    cell.fill = PatternFill(
                        start_color='FCA5A5',
                        end_color='FCA5A5',
                        fill_type='solid'
                    )

                elif value == 'ALTA':

                    cell.fill = PatternFill(
                        start_color='FDBA74',
                        end_color='FDBA74',
                        fill_type='solid'
                    )

                elif value == 'MEDIA':

                    cell.fill = PatternFill(
                        start_color='FDE68A',
                        end_color='FDE68A',
                        fill_type='solid'
                    )

                elif value == 'BAJA':

                    cell.fill = PatternFill(
                        start_color='86EFAC',
                        end_color='86EFAC',
                        fill_type='solid'
                    )

        current_row += 1

    # =================================================
    # TAMAÑO COLUMNAS
    # =================================================

    columnas = {

        'A': 10,
        'B': 22,
        'C': 20,
        'D': 20,
        'E': 15,
        'F': 15,
        'G': 25
    }

    for col, width in columnas.items():

        ws.column_dimensions[col].width = width

    # =================================================
    # GENERAR ARCHIVO
    # =================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        download_name=
            f'REPORTE_ALERTAS_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',

        as_attachment=True,

        mimetype=
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# =====================================================
# SEMÁFORO ALERTAS
# =====================================================

@reportes_bp.route(
    '/semaforo-alertas',
    methods=['GET']
)
def semaforo_alertas():

    criticas = Alerta.query.filter_by(
        prioridad='CRITICA',
        estado='ACTIVA'
    ).count()

    altas = Alerta.query.filter_by(
        prioridad='ALTA',
        estado='ACTIVA'
    ).count()

    medias = Alerta.query.filter_by(
        prioridad='MEDIA',
        estado='ACTIVA'
    ).count()

    bajas = Alerta.query.filter_by(
        prioridad='BAJA',
        estado='ACTIVA'
    ).count()

    total = (
        criticas
        + altas
        + medias
        + bajas
    )

    return jsonify({

        'total': total,

        'criticas': criticas,

        'altas': altas,

        'medias': medias,

        'bajas': bajas
    })

# =====================================================
# REPORTE MANTENIMIENTOS
# =====================================================
@reportes_bp.route(
    '/mantenimientos',
    methods=['GET']
)
def reporte_mantenimientos():

    inicio, fin = obtener_rango_fechas()

    query = (
        Mantenimiento.query
        .filter(
            Mantenimiento.fecha >= inicio,
            Mantenimiento.fecha <= fin
        )
    )

    vehiculo_id = request.args.get(
        'vehiculo_id'
    )

    # =========================================
    # FILTRO VEHÍCULO
    # =========================================

    if vehiculo_id:

        query = query.filter(
            Mantenimiento.vehiculo_id == vehiculo_id
        )

    mantenimientos = query.all()

    data = []

    for m in mantenimientos:

        vehiculo = Vehiculo.query.get(
            m.vehiculo_id
        )

        data.append({

            'id': m.id,

            'vehiculo':
                vehiculo.placa
                if vehiculo else None,

            'tipo':
                m.tipo,

            'fecha':
                str(m.fecha),

            'km':
                m.km,

            'observaciones':
                m.observaciones

        })

    return jsonify(data)


# =====================================================
# FORMATO PROFESIONAL MANTENIMIENTO VEHÍCULO
# =====================================================

@reportes_bp.route(
    '/mantenimiento-formato/<int:vehiculo_id>',
    methods=['GET']
)
def exportar_formato_mantenimiento(vehiculo_id):

    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font,
        PatternFill,
        Border,
        Side,
        Alignment
    )

    # =================================================
    # VEHÍCULO
    # =================================================

    vehiculo = Vehiculo.query.get_or_404(
        vehiculo_id
    )

    # =================================================
    # MANTENIMIENTOS
    # =================================================

    mantenimientos = (

        Mantenimiento.query

        .filter(
            Mantenimiento.vehiculo_id == vehiculo_id
        )

        .order_by(
            Mantenimiento.fecha.desc()
        )

        .all()
    )

    # =================================================
    # EXCEL
    # =================================================

    wb = Workbook()

    ws = wb.active

    ws.title = 'R. MANTENIMIENTO'

    # =================================================
    # OCULTAR GRID
    # =================================================

    ws.sheet_view.showGridLines = False

    # =================================================
    # ANCHO COLUMNAS
    # =================================================

    columnas = {

        'A': 18,
        'B': 18,
        'C': 18,
        'D': 18,
        'E': 16,
        'F': 16,
        'G': 18,
        'H': 12,
        'I': 12,
        'J': 12

    }

    for col, width in columnas.items():

        ws.column_dimensions[col].width = width

    # =================================================
    # ESTILOS
    # =================================================

    azul_oscuro = PatternFill(
        start_color='1F4E78',
        end_color='1F4E78',
        fill_type='solid'
    )

    azul_claro = PatternFill(
        start_color='D9EAF7',
        end_color='D9EAF7',
        fill_type='solid'
    )

    azul_header = PatternFill(
        start_color='8DB4E2',
        end_color='8DB4E2',
        fill_type='solid'
    )

    gris = PatternFill(
        start_color='F2F2F2',
        end_color='F2F2F2',
        fill_type='solid'
    )

    thin = Side(
        border_style='thin',
        color='BFBFBF'
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )

    left = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )

    titulo_font = Font(
        bold=True,
        color='FFFFFF',
        size=14
    )

    subtitulo_font = Font(
        bold=True,
        color='FFFFFF',
        size=11
    )

    bold = Font(
        bold=True,
        size=10,
        color='1F1F1F'
    )

    normal = Font(
        size=10,
        color='333333'
    )

    white_bold = Font(
        bold=True,
        size=10,
        color='FFFFFF'
    )

    # =================================================
    # ALTURA FILAS
    # =================================================

    for i in range(1, 60):

        ws.row_dimensions[i].height = 28

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30

    # =================================================
    # HEADER PRINCIPAL
    # =================================================

    # =================================================
# LOGO EMPRESA
# =================================================

    ws.merge_cells('A1:B3')

# Bordes y fondo para el área del logo
    for row in ws['A1:B3']:

        for cell in row:

            cell.fill = azul_oscuro
            cell.border = border

    # Ruta logo
    logo = Image('static/logo_transmena.jpg')

    # Tamaño logo
    logo.width = 260
    logo.height = 128

    # Posición
    ws.add_image(logo, 'A1')
    ws.merge_cells('C1:G2')

    ws['C1'] = 'GESTIÓN DIRECCIÓN DE PROYECTOS'

    ws['C1'].fill = azul_oscuro
    ws['C1'].font = titulo_font
    ws['C1'].alignment = center
    ws['C1'].border = border

    ws.merge_cells('C3:G3')

    ws['C3'] = 'REPORTE DE MANTENIMIENTO VEHICULAR'

    ws['C3'].fill = azul_claro
    ws['C3'].font = Font(
        bold=True,
        size=12,
        color='1F1F1F'
    )

    ws['C3'].alignment = center
    ws['C3'].border = border

    # =================================================
    # VERSIONES
    # =================================================

    info_header = [

        ('H1:I1', 'VERSIÓN: 006'),
        ('H2:I2', 'CÓDIGO: PROY-R-022'),
        ('H3:I3', 'PÁGINA: 1 DE 1')

    ]

    for rango, texto in info_header:

        ws.merge_cells(rango)

        cell = rango.split(':')[0]

        ws[cell] = texto

        ws[cell].fill = azul_oscuro
        ws[cell].font = white_bold
        ws[cell].alignment = center
        ws[cell].border = border

    # =================================================
    # SECCIÓN DATOS VEHÍCULO
    # =================================================

    ws.merge_cells('A5:I5')

    ws['A5'] = 'INFORMACIÓN GENERAL DEL VEHÍCULO'

    ws['A5'].fill = azul_oscuro
    ws['A5'].font = subtitulo_font
    ws['A5'].alignment = center
    ws['A5'].border = border

    tipo_vehiculo = ''

    if vehiculo.tipo_vehiculo:

        tipo_vehiculo = getattr(
            vehiculo.tipo_vehiculo,
            'nombre',
            str(vehiculo.tipo_vehiculo)
        )

    datos = [

        ('CLASE VEHÍCULO', tipo_vehiculo),
        ('MODELO', str(vehiculo.modelo or '')),

        ('PLACA', str(vehiculo.placa or '')),
        ('MARCA', str(vehiculo.marca or '')),

        ('KILÓMETRAJE', str(vehiculo.km_actual or '')),
        ('ESTADO', str(getattr(vehiculo, 'estado', 'ACTIVO')))

    ]

    posiciones = [

        ('A7:B7', 'C7:E7'),
        ('F7:G7', 'H7:I7'),

        ('A9:B9', 'C9:E9'),
        ('F9:G9', 'H9:I9'),

        ('A11:B11', 'C11:E11'),
        ('F11:G11', 'H11:I11')

    ]

    for i, (label_pos, value_pos) in enumerate(posiciones):

        label, value = datos[i]

        # LABEL
        ws.merge_cells(label_pos)

        label_cell = label_pos.split(':')[0]

        ws[label_cell] = label

        ws[label_cell].fill = azul_claro
        ws[label_cell].font = bold
        ws[label_cell].alignment = center
        ws[label_cell].border = border

        # VALUE
        ws.merge_cells(value_pos)

        value_cell = value_pos.split(':')[0]

        ws[value_cell] = value

        ws[value_cell].fill = gris
        ws[value_cell].font = normal
        ws[value_cell].alignment = center
        ws[value_cell].border = border

    # =================================================
    # HEADER TABLA
    # =================================================

    headers = [

        'DESCRIPCIÓN DETALLADA DEL\nMANTENIMIENTO',

        'INSUMOS Y\nREPUESTOS',

        'ENTIDAD Y/O\nRESPONSABLE',

        'MANTENIMIENTO\nPREVENTIVO',

        'MANTENIMIENTO\nCORRECTIVO',
        'SOPORTE'

    ]

    merges = [
        'A14:D15',
        'E14:F15',
        'G14:G15',
        'H14:H15',
        'I14:I15',
        'J14:J15'
    ]

    for i, merge in enumerate(merges):

        ws.merge_cells(merge)

        cell = merge.split(':')[0]

        ws[cell] = headers[i]

        ws[cell].fill = azul_header

        ws[cell].font = Font(
            bold=True,
            size=9
        )

        ws[cell].alignment = center

        ws[cell].border = border

    ws.row_dimensions[14].height = 38
    ws.row_dimensions[15].height = 38

    # =================================================
    # TABLA MANTENIMIENTOS
    # =================================================

    fila_actual = 16

    for m in mantenimientos[:18]:

        ws.merge_cells(
            f'A{fila_actual}:D{fila_actual}'
        )

        ws[f'A{fila_actual}'] = (
            getattr(m.plan_item, 'nombre', '')
        )

        ws.merge_cells(
            f'E{fila_actual}:F{fila_actual}'
        )

        ws[f'E{fila_actual}'] = (
            str(getattr(m, 'proveedor', '') or '')
        )

        ws[f'G{fila_actual}'] = (
            str(getattr(m, 'responsable', '') or '')
        )
        if m.soporte:
            ruta_imagen = os.path.join(
                os.getcwd(),
                m.soporte
            )

            if os.path.exists(ruta_imagen):

                img = Image(ruta_imagen)

                img.width = 180
                img.height = 140

        # Columna J
            ws.add_image(
                img,
                f'J{fila_actual}'
            )

        if m.tipo == 'PREVENTIVO':

            ws[f'H{fila_actual}'] = '✔'

        else:

            ws[f'I{fila_actual}'] = '✔'

        for col in range(1, 10):

            cell = ws.cell(
                row=fila_actual,
                column=col
            )

            cell.border = border

            cell.alignment = Alignment(
                wrap_text=True,
                vertical='center',
                horizontal='center'
            )

            cell.font = normal

        ws[f'A{fila_actual}'].alignment = left

        if fila_actual % 2 == 0:

            for col in range(1, 10):

                ws.cell(
                    row=fila_actual,
                    column=col
                ).fill = gris

        ws.row_dimensions[fila_actual].height = 100
        fila_actual += 1

    # =================================================
    # FILAS VACÍAS
    # =================================================

    while fila_actual <= 34:

        for col in range(1, 10):

            cell = ws.cell(
                row=fila_actual,
                column=col
            )

            cell.border = border

            cell.alignment = center

            if fila_actual % 2 == 0:

                cell.fill = gris

        ws.row_dimensions[fila_actual].height = 30

        fila_actual += 1

    # =================================================
    # TABLA CONVENCIONES
    # =================================================

    ws.merge_cells('A36:I36')

    ws['A36'] = (
        'TABLA DE CONVENCIONES - SISTEMAS DE MANTENIMIENTO'
    )

    ws['A36'].fill = azul_oscuro

    ws['A36'].font = subtitulo_font

    ws['A36'].alignment = center

    ws['A36'].border = border

    convenciones = [

        ('SISTEMA DE LUBRICACIÓN', 'SL'),
        ('SISTEMA DE COMBUSTIBLE', 'SC'),
        ('SISTEMA ELÉCTRICO', 'SEL'),
        ('SISTEMA DE FRENOS', 'SF'),
        ('SISTEMA DE TRANSMISIÓN', 'ST'),

    ]

    fila_conv = 37

    for nombre, sigla in convenciones:

        ws.merge_cells(
            f'A{fila_conv}:B{fila_conv}'
        )

        ws[f'A{fila_conv}'] = nombre

        ws[f'C{fila_conv}'] = sigla

        ws[f'A{fila_conv}'].border = border
        ws[f'C{fila_conv}'].border = border

        ws[f'A{fila_conv}'].alignment = center
        ws[f'C{fila_conv}'].alignment = center

        ws[f'A{fila_conv}'].fill = azul_claro
        ws[f'C{fila_conv}'].fill = gris

        fila_conv += 1

    convenciones2 = [

        ('SISTEMA DE DIRECCIÓN', 'SD'),
        ('SISTEMA DE MOTOR', 'SM'),
        ('SISTEMA DE SUSPENSIÓN', 'SS'),
        ('SISTEMA DE ESCAPE', 'SES'),
        ('SISTEMA DE LLANTAS', 'SLL'),

    ]

    fila_conv = 37

    for nombre, sigla in convenciones2:

        ws.merge_cells(
            f'F{fila_conv}:G{fila_conv}'
        )

        ws[f'F{fila_conv}'] = nombre

        ws[f'H{fila_conv}'] = sigla

        ws[f'F{fila_conv}'].border = border
        ws[f'H{fila_conv}'].border = border

        ws[f'F{fila_conv}'].alignment = center
        ws[f'H{fila_conv}'].alignment = center

        ws[f'F{fila_conv}'].fill = azul_claro
        ws[f'H{fila_conv}'].fill = gris

        fila_conv += 1

    # =================================================
    # EXPORTAR
    # =================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        download_name=(
            f'FORMATO_MANTENIMIENTO_'
            f'{vehiculo.placa}.xlsx'
        ),

        as_attachment=True,

        mimetype=(
            'application/vnd.openxmlformats-'
            'officedocument.spreadsheetml.sheet'
        )
    )
    
    
@reportes_bp.route(
    '/alertas-formato/<int:vehiculo_id>',
    methods=['GET']
)
def exportar_formato_alertas(vehiculo_id):

    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import (
        Font,
        PatternFill,
        Border,
        Side,
        Alignment
    )

    vehiculo = Vehiculo.query.get_or_404(vehiculo_id)

    alertas = (
        Alerta.query
        .filter(Alerta.vehiculo_id == vehiculo_id)
        .order_by(Alerta.created_at.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active

    ws.title = 'R. ALERTAS'
    ws.sheet_view.showGridLines = False

    # =================================================
    # COLUMNAS
    # =================================================

    columnas = {
        'A': 18,
        'B': 18,
        'C': 18,
        'D': 18,
        'E': 16,
        'F': 16,
        'G': 18,
        'H': 12,
        'I': 12
    }

    for col, width in columnas.items():
        ws.column_dimensions[col].width = width

    # =================================================
    # ESTILOS
    # =================================================

    azul_oscuro = PatternFill(start_color='1F4E78', fill_type='solid')
    azul_claro = PatternFill(start_color='D9EAF7', fill_type='solid')
    azul_header = PatternFill(start_color='8DB4E2', fill_type='solid')
    gris = PatternFill(start_color='F2F2F2', fill_type='solid')

    thin = Side(border_style='thin', color='BFBFBF')

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )

    left = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )

    titulo_font = Font(bold=True, color='FFFFFF', size=14)
    subtitulo_font = Font(bold=True, color='FFFFFF', size=11)
    normal = Font(size=10, color='333333')
    white_bold = Font(bold=True, color='FFFFFF', size=10)

    # =================================================
    # ALTURA FILAS
    # =================================================

    for i in range(1, 60):
        ws.row_dimensions[i].height = 28

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30

    # =================================================
    # HEADER (NO TOCADO)
    # =================================================

    ws.merge_cells('A1:B3')

    for row in ws['A1:B3']:
        for cell in row:
            cell.fill = azul_oscuro
            cell.border = border

    logo = Image('static/logo_transmena.jpg')
    logo.width = 260
    logo.height = 128
    ws.add_image(logo, 'A1')

    ws.merge_cells('C1:G2')

    ws['C1'] = 'GESTIÓN DIRECCIÓN DE PROYECTOS'
    ws['C1'].fill = azul_oscuro
    ws['C1'].font = titulo_font
    ws['C1'].alignment = center
    ws['C1'].border = border

    ws.merge_cells('C3:G3')

    ws['C3'] = 'REPORTE DE ALERTAS VEHICULARES'
    ws['C3'].fill = azul_claro
    ws['C3'].font = Font(bold=True, size=12, color='1F1F1F')
    ws['C3'].alignment = center
    ws['C3'].border = border

    # =================================================
    # INFO HEADER
    # =================================================

    info_header = [
        ('H1:I1', 'VERSIÓN: 001'),
        ('H2:I2', 'CÓDIGO: ALERT-R-001'),
        ('H3:I3', 'PÁGINA: 1 DE 1')
    ]

    for rango, texto in info_header:
        ws.merge_cells(rango)
        cell = rango.split(':')[0]

        ws[cell] = texto
        ws[cell].fill = azul_oscuro
        ws[cell].font = white_bold
        ws[cell].alignment = center
        ws[cell].border = border

    # =================================================
    # VEHÍCULO
    # =================================================

    ws.merge_cells('A5:I5')

    ws['A5'] = 'INFORMACIÓN GENERAL DEL VEHÍCULO'
    ws['A5'].fill = azul_oscuro
    ws['A5'].font = subtitulo_font
    ws['A5'].alignment = center
    ws['A5'].border = border

    datos = [
        ('CLASE VEHÍCULO', getattr(vehiculo.tipo_vehiculo, 'nombre', '')),
        ('MODELO', vehiculo.modelo or ''),
        ('PLACA', vehiculo.placa or ''),
        ('MARCA', vehiculo.marca or ''),
        ('KM ACTUAL', vehiculo.km_actual or ''),
        ('ESTADO', getattr(vehiculo, 'estado', 'ACTIVO'))
    ]

    posiciones = [
        ('A7:B7', 'C7:E7'),
        ('F7:G7', 'H7:I7'),
        ('A9:B9', 'C9:E9'),
        ('F9:G9', 'H9:I9'),
        ('A11:B11', 'C11:E11'),
        ('F11:G11', 'H11:I11')
    ]

    for i, (label_pos, value_pos) in enumerate(posiciones):

        label, value = datos[i]

        ws.merge_cells(label_pos)
        lc = label_pos.split(':')[0]

        ws[lc] = label
        ws[lc].fill = azul_claro
        ws[lc].font = Font(bold=True, size=10)
        ws[lc].alignment = center
        ws[lc].border = border

        ws.merge_cells(value_pos)
        vc = value_pos.split(':')[0]

        ws[vc] = value
        ws[vc].fill = gris
        ws[vc].alignment = center
        ws[vc].border = border

    # =================================================
    # HEADER TABLA
    # =================================================

    headers = ['TIPO', 'CATEGORÍA', 'PRIORIDAD', 'ESTADO', 'MENSAJE']
    merges = ['A14:B15', 'C14:C15', 'D14:D15', 'E14:E15', 'F14:I15']

    for i, merge in enumerate(merges):

        ws.merge_cells(merge)
        cell = merge.split(':')[0]

        ws[cell] = headers[i]
        ws[cell].fill = azul_header
        ws[cell].font = Font(bold=True, size=10)
        ws[cell].alignment = center
        ws[cell].border = border

    ws.row_dimensions[14].height = 38
    ws.row_dimensions[15].height = 38

    # =================================================
    # COLORES POR ESTADO / PRIORIDAD
    # =================================================

    def color(valor):
        if valor == 'CRITICA':
            return PatternFill(start_color='FCA5A5', fill_type='solid')
        if valor == 'ALTA':
            return PatternFill(start_color='FDBA74', fill_type='solid')
        if valor == 'MEDIA':
            return PatternFill(start_color='FDE68A', fill_type='solid')
        if valor == 'BAJA':
            return PatternFill(start_color='86EFAC', fill_type='solid')
        return None

    # =================================================
    # TABLA
    # =================================================

    fila = 16

    for a in alertas:

        ws.merge_cells(f'A{fila}:B{fila}')
        ws.merge_cells(f'F{fila}:I{fila}')  # MENSAJE BIEN COMBINADO

        ws[f'A{fila}'] = a.tipo
        ws[f'C{fila}'] = a.categoria
        ws[f'D{fila}'] = a.prioridad
        ws[f'E{fila}'] = a.estado
        ws[f'F{fila}'] = a.mensaje

        # base estilo
        for col in range(1, 10):
            c = ws.cell(row=fila, column=col)
            c.border = border
            c.alignment = center
            c.font = normal

        ws[f'A{fila}'].alignment = left
        ws[f'F{fila}'].alignment = left

        # colores
        if color(a.prioridad):
            ws[f'D{fila}'].fill = color(a.prioridad)

        if color(a.estado):
            ws[f'E{fila}'].fill = color(a.estado)

        # zebra
        if fila % 2 == 0:
            for col in range(1, 10):
                ws.cell(row=fila, column=col).fill = gris

        fila += 1

    # =================================================
    # BORDE EXTERIOR (CUADRO FINAL)
    # =================================================

    for r in range(14, fila):
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = border

    # =================================================
    # EXPORTAR
    # =================================================

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name=f'FORMATO_ALERTAS_{vehiculo.placa}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )