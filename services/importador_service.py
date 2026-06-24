from openpyxl import load_workbook

from models import (
    db,

    Vehiculo,
    TipoVehiculo,

    VehiculoDocumento,
    DocumentoTipo,

    VehiculoCampoValor,
    TipoVehiculoCampo
)

# =====================================================
# IMPORTAR EXCEL
# =====================================================

def importar_excel_service(archivo):

    wb = load_workbook(archivo)

    errores = []

    resultado = {
        "vehiculos": 0,
        "documentos": 0,
        "campos_especiales": 0,
        "errores": errores
    }

    # =====================================================
    # VEHICULOS
    # =====================================================

    if 'VEHICULOS' in wb.sheetnames:

        ws = wb['VEHICULOS']

        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            try:

                placa = row[0]

                if not placa:
                    continue

                tipo_nombre = row[15]

                tipo_vehiculo = (
                    TipoVehiculo.query
                    .filter_by(
                        nombre=tipo_nombre
                    )
                    .first()
                )

                if not tipo_vehiculo:

                    errores.append(
                        f'Tipo vehículo inválido: {tipo_nombre}'
                    )

                    continue

                vehiculo = (
                    Vehiculo.query
                    .filter_by(
                        placa=placa
                    )
                    .first()
                )

                # =====================================
                # CREAR
                # =====================================
                if not vehiculo:

                    vehiculo = Vehiculo(
                        placa=placa
                    )

                    db.session.add(
                        vehiculo
                    )

                # =====================================
                # DATOS
                # =====================================
                vehiculo.marca = row[1]
                vehiculo.linea = row[2]
                vehiculo.modelo = row[3]
                vehiculo.color = row[4]
                vehiculo.vin = row[5]
                vehiculo.numero_chasis = row[6]
                vehiculo.numero_motor = row[7]
                vehiculo.propietario = row[8]
                vehiculo.cc_propietario = row[9]
                vehiculo.conductor = row[10]
                vehiculo.cc_conductor = row[11]
                vehiculo.servicio = row[12]
                vehiculo.km_actual = row[13] or 0
                vehiculo.gps_id = row[14]

                vehiculo.tipo_vehiculo_id = (
                    tipo_vehiculo.id
                )

                vehiculo.tipo_operacion = row[16]
                vehiculo.estado = row[17]

                vehiculo.tiene_gps = (
                    str(row[18]).upper() == 'SI'
                )

                vehiculo.notas = row[19]

                resultado["vehiculos"] += 1

            except Exception as e:

                errores.append(
                    f'VEHICULOS {placa}: {str(e)}'
                )

    db.session.commit()

    # =====================================================
    # DOCUMENTOS
    # =====================================================

    if 'DOCUMENTOS' in wb.sheetnames:

        ws = wb['DOCUMENTOS']

        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            try:

                placa = row[0]

                tipo_doc_nombre = row[1]

                vehiculo = (
                    Vehiculo.query
                    .filter_by(
                        placa=placa
                    )
                    .first()
                )

                if not vehiculo:

                    errores.append(
                        f'Vehículo no encontrado: {placa}'
                    )

                    continue

                tipo_doc = (
                    DocumentoTipo.query
                    .filter_by(
                        nombre=tipo_doc_nombre
                    )
                    .first()
                )

                if not tipo_doc:

                    errores.append(
                        f'Documento inválido: {tipo_doc_nombre}'
                    )

                    continue

                documento = VehiculoDocumento(

                    vehiculo_id=vehiculo.id,

                    documento_tipo_id=tipo_doc.id,

                    numero=row[2],

                    fecha_expedicion=row[3],

                    fecha_vencimiento=row[4],

                    archivo_url=row[5]
                )

                db.session.add(
                    documento
                )

                resultado["documentos"] += 1

            except Exception as e:

                errores.append(
                    f'DOCUMENTOS {placa}: {str(e)}'
                )

    db.session.commit()

    # =====================================================
    # CAMPOS ESPECIALES
    # =====================================================

    if 'CAMPOS_ESPECIALES' in wb.sheetnames:

        ws = wb['CAMPOS_ESPECIALES']

        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            try:

                placa = row[0]

                campo_nombre = row[1]

                valor = row[2]

                vehiculo = (
                    Vehiculo.query
                    .filter_by(
                        placa=placa
                    )
                    .first()
                )

                if not vehiculo:

                    errores.append(
                        f'Vehículo no encontrado: {placa}'
                    )

                    continue

                campo = (
                    TipoVehiculoCampo.query
                    .filter_by(
                        nombre_campo=campo_nombre
                    )
                    .first()
                )

                if not campo:

                    errores.append(
                        f'Campo inválido: {campo_nombre}'
                    )

                    continue

                existente = (
                    VehiculoCampoValor.query
                    .filter_by(
                        vehiculo_id=vehiculo.id,
                        campo_id=campo.id
                    )
                    .first()
                )

                # =====================================
                # ACTUALIZAR
                # =====================================
                if existente:

                    existente.valor = str(
                        valor
                    )

                # =====================================
                # CREAR
                # =====================================
                else:

                    nuevo = VehiculoCampoValor(

                        vehiculo_id=vehiculo.id,

                        campo_id=campo.id,

                        valor=str(valor)
                    )

                    db.session.add(
                        nuevo
                    )

                resultado["campos_especiales"] += 1

            except Exception as e:

                errores.append(
                    f'CAMPOS {placa}: {str(e)}'
                )

    db.session.commit()

    return resultado