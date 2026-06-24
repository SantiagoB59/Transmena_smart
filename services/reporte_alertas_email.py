from io import BytesIO
from datetime import datetime, date

from flask_mail import Message
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from extensions import mail
from models import Alerta

import os


# =====================================================
# ENVIAR CONSOLIDADO DIARIO DE ALERTAS
# =====================================================
def enviar_reporte_diario_alertas(app):

    with app.app_context():

        print("📨 Ejecutando consolidado diario...")

        hoy = date.today()

        alertas = (
            Alerta.query
            .filter(
                Alerta.created_at >= datetime.combine(
                    hoy,
                    datetime.min.time()
                )
            )
            .order_by(
                Alerta.created_at.desc()
            )
            .all()
        )

        archivo_excel = generar_excel_alertas(alertas)

        destinatario = os.getenv(
            "ALERTA_EMAIL",
            "ferney.bonilla59@gmail.com"
        )

        total = len(alertas)

        criticas = len([
            a for a in alertas
            if a.prioridad == 'CRITICA'
        ])

        altas = len([
            a for a in alertas
            if a.prioridad == 'ALTA'
        ])

        medias = len([
            a for a in alertas
            if a.prioridad == 'MEDIA'
        ])

        bajas = len([
            a for a in alertas
            if a.prioridad == 'BAJA'
        ])

        html = f"""
        <html>
        <body style="
            font-family: Arial, sans-serif;
            background:#f4f6f9;
            padding:30px;
        ">

            <div style="
                max-width:900px;
                margin:auto;
                background:white;
                border-radius:12px;
                overflow:hidden;
                box-shadow:0 2px 10px rgba(0,0,0,.1);
            ">

                <div style="
                    background:#0f172a;
                    color:white;
                    padding:30px;
                    text-align:center;
                ">
                    <h1>🚛 TRANSMENA SMART</h1>
                    <h2>Consolidado Diario de Alertas</h2>
                    <p>{hoy.strftime('%d/%m/%Y')}</p>
                </div>

                <div style="padding:30px;">

                    <h3>Resumen Ejecutivo</h3>

                    <table
                        width="100%"
                        cellpadding="10"
                        cellspacing="0"
                        style="
                            border-collapse:collapse;
                            text-align:center;
                        "
                    >
                        <tr>
                            <td style="background:#fee2e2;">
                                <b>Críticas</b><br>{criticas}
                            </td>

                            <td style="background:#ffedd5;">
                                <b>Altas</b><br>{altas}
                            </td>

                            <td style="background:#fef9c3;">
                                <b>Medias</b><br>{medias}
                            </td>

                            <td style="background:#dcfce7;">
                                <b>Bajas</b><br>{bajas}
                            </td>
                        </tr>
                    </table>

                    <br>

                    <h2>
                        Total Alertas Registradas:
                        {total}
                    </h2>

                    <p>
                        Se adjunta el archivo Excel con el detalle
                        completo de las alertas registradas durante
                        la jornada.
                    </p>

                </div>

                <div style="
                    background:#f8fafc;
                    padding:20px;
                    text-align:center;
                    color:#64748b;
                    font-size:12px;
                ">
                    Sistema de Monitoreo Vehicular
                    TRANSMENA SMART
                </div>

            </div>

        </body>
        </html>
        """

        msg = Message(
            subject=f"📊 Consolidado Diario de Alertas - {hoy}",
            recipients=[destinatario]
        )

        msg.html = html

        msg.attach(
            filename=f"ALERTAS_{hoy}.xlsx",
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            data=archivo_excel.getvalue()
        )

        print(f"📧 Enviando a: {destinatario}")

        mail.send(msg)

        print(
            f"✅ Consolidado diario enviado "
            f"({total} alertas)"
        )
# =====================================================
# EXCEL PROFESIONAL
# =====================================================

def generar_excel_alertas(alertas):

    wb = Workbook()

    ws = wb.active

    ws.title = "CONSOLIDADO ALERTAS"

    ws.sheet_view.showGridLines = False

    columnas = {
        'A': 18,
        'B': 18,
        'C': 18,
        'D': 18,
        'E': 15,
        'F': 15,
        'G': 50,
        'H': 25,
        'I': 25
    }

    for col, width in columnas.items():
        ws.column_dimensions[col].width = width

    azul_oscuro = PatternFill(
        start_color='1F4E78',
        fill_type='solid'
    )

    azul_claro = PatternFill(
        start_color='D9EAF7',
        fill_type='solid'
    )

    azul_header = PatternFill(
        start_color='8DB4E2',
        fill_type='solid'
    )

    gris = PatternFill(
        start_color='F2F2F2',
        fill_type='solid'
    )

    thin = Side(
        border_style='thin',
        color='BFBFBF'
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )

    left = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )

    white_bold = Font(
        bold=True,
        color='FFFFFF'
    )

    # =================================================
    # LOGO
    # =================================================

    ws.merge_cells('A1:B3')

    for row in ws['A1:B3']:
        for cell in row:
            cell.fill = azul_oscuro
            cell.border = border

    try:

        logo = Image(
            'static/logo_transmena.jpg'
        )

        logo.width = 250
        logo.height = 120

        ws.add_image(
            logo,
            'A1'
        )

    except Exception:
        pass

    # =================================================
    # TITULO
    # =================================================

    ws.merge_cells('C1:G2')

    ws['C1'] = (
        'GESTIÓN DIRECCIÓN DE PROYECTOS'
    )

    ws['C1'].fill = azul_oscuro
    ws['C1'].font = Font(
        bold=True,
        color='FFFFFF',
        size=14
    )
    ws['C1'].alignment = center
    ws['C1'].border = border

    ws.merge_cells('C3:G3')

    ws['C3'] = (
        'CONSOLIDADO GENERAL DE ALERTAS'
    )

    ws['C3'].fill = azul_claro
    ws['C3'].alignment = center
    ws['C3'].border = border

    info = [
        ('H1:I1', 'VERSIÓN: 001'),
        ('H2:I2', 'CÓDIGO: ALERT-R-001'),
        ('H3:I3', 'PÁGINA: 1 DE 1')
    ]

    for rango, texto in info:

        ws.merge_cells(rango)

        c = rango.split(':')[0]

        ws[c] = texto

        ws[c].fill = azul_oscuro

        ws[c].font = white_bold

        ws[c].alignment = center

        ws[c].border = border

    # =================================================
    # RESUMEN
    # =================================================

    ws.merge_cells('A5:I5')

    ws['A5'] = (
        f'REPORTE GENERADO '
        f'{datetime.now().strftime("%d/%m/%Y %H:%M")}'
    )

    ws['A5'].fill = azul_oscuro
    ws['A5'].font = white_bold
    ws['A5'].alignment = center

    # =================================================
    # HEADERS
    # =================================================

    headers = [
        'FECHA',
        'TIPO',
        'CATEGORIA',
        'PRIORIDAD',
        'ESTADO',
        'VEHICULO',
        'MENSAJE',
        'ORIGEN',
        'EVENTO'
    ]

    fila = 8

    for col, texto in enumerate(headers, start=1):

        cell = ws.cell(
            row=fila,
            column=col
        )

        cell.value = texto

        cell.fill = azul_header

        cell.font = Font(
            bold=True
        )

        cell.alignment = center

        cell.border = border

    fila += 1

    # =================================================
    # COLORES
    # =================================================

    def color_prioridad(valor):

        if valor == 'CRITICA':
            return PatternFill(
                start_color='FCA5A5',
                fill_type='solid'
            )

        if valor == 'ALTA':
            return PatternFill(
                start_color='FDBA74',
                fill_type='solid'
            )

        if valor == 'MEDIA':
            return PatternFill(
                start_color='FDE68A',
                fill_type='solid'
            )

        if valor == 'BAJA':
            return PatternFill(
                start_color='86EFAC',
                fill_type='solid'
            )

        return None

    # =================================================
    # DATOS
    # =================================================

    for alerta in alertas:

        ws.cell(
            fila,
            1,
            alerta.created_at.strftime(
                '%d/%m/%Y %H:%M'
            )
            if alerta.created_at
            else ''
        )

        ws.cell(fila, 2, alerta.tipo)
        ws.cell(fila, 3, alerta.categoria)
        ws.cell(fila, 4, alerta.prioridad)
        ws.cell(fila, 5, alerta.estado)

        ws.cell(
            fila,
            6,
            alerta.vehiculo.placa
            if alerta.vehiculo
            else ''
        )

        ws.cell(
            fila,
            7,
            alerta.mensaje
        )

        ws.cell(
            fila,
            8,
            alerta.origen
        )

        ws.cell(
            fila,
            9,
            alerta.titulo
        )

        for c in range(1, 10):

            cell = ws.cell(fila, c)

            cell.border = border

            cell.alignment = left

            if fila % 2 == 0:
                cell.fill = gris

        color = color_prioridad(
            alerta.prioridad
        )

        if color:
            ws.cell(
                fila,
                4
            ).fill = color

        fila += 1

    # =================================================
    # EXPORTAR
    # =================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output